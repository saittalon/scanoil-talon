import csv
import io
import json
import os
import posixpath
import zipfile
from datetime import date, datetime

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client

from models import (
    User, Client, Contract, ContractFile, Balance, Talon,
    AGZS, WebAppToken, BotSession, TalonRedemption, Shift,
    AuditLog, RateLimitEvent,
)

EXPORT_MODELS = [
    (User, 'users'),
    (Client, 'clients'),
    (Contract, 'contracts'),
    (ContractFile, 'contract_files'),
    (Balance, 'balances'),
    (Talon, 'talons'),
    (AGZS, 'agzs'),
    (WebAppToken, 'webapp_tokens'),
    (BotSession, 'bot_sessions'),
    (TalonRedemption, 'talon_redemptions'),
    (Shift, 'shifts'),
    (AuditLog, 'audit_logs'),
    (RateLimitEvent, 'rate_limit_events'),
]

HEADER_FILL = PatternFill('solid', fgColor='0F2747')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(size=14, bold=True)
SUBTITLE_FONT = Font(color='666666')


def _serialize(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_to_dict(obj):
    data = {}
    for column in obj.__table__.columns:
        data[column.name] = _serialize(getattr(obj, column.name))
    return data


def _csv_bytes(rows, fieldnames):
    stream = io.StringIO()
    writer = csv.DictWriter(stream, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return stream.getvalue().encode('utf-8-sig')


def _get_supabase_client():
    supabase_url = os.getenv('SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
    if not supabase_url or not supabase_key:
        return None
    return create_client(supabase_url, supabase_key)


def collect_backup_rows():
    data = {}
    for model, folder in EXPORT_MODELS:
        rows = [_row_to_dict(item) for item in model.query.order_by(model.id.asc()).all()]
        fieldnames = list(rows[0].keys()) if rows else [c.name for c in model.__table__.columns]
        data[folder] = {
            'rows': rows,
            'fieldnames': fieldnames,
        }
    return data


def _sheet_title(name: str) -> str:
    clean = name.replace('_', ' ').title()
    return clean[:31]


def build_backup_excel_bytes(data=None, timestamp: str | None = None):
    data = data or collect_backup_rows()
    timestamp = timestamp or datetime.utcnow().strftime('%Y%m%d-%H%M%S')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Overview'
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Scanoil Backup Overview'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = 'Backup generated automatically for convenient Excel viewing.'
    ws['A2'].font = SUBTITLE_FONT
    ws['A4'] = 'Generated UTC'
    ws['B4'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    ws['A5'] = 'Backup stamp'
    ws['B5'] = timestamp
    ws['A7'] = 'Sheet'
    ws['B7'] = 'Rows'
    for cell in ws[7]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    row_idx = 8
    for folder, payload in data.items():
        ws.cell(row=row_idx, column=1, value=folder)
        ws.cell(row=row_idx, column=2, value=len(payload['rows']))
        row_idx += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.freeze_panes = 'A8'
    ws.auto_filter.ref = f'A7:B{max(7, row_idx - 1)}'

    for folder, payload in data.items():
        sheet = wb.create_sheet(_sheet_title(folder))
        sheet.sheet_view.showGridLines = False
        sheet['A1'] = f'{folder} backup export'
        sheet['A1'].font = TITLE_FONT
        sheet['A2'] = f'Rows: {len(payload["rows"])}'
        sheet['A2'].font = SUBTITLE_FONT

        fieldnames = payload['fieldnames']
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=4, column=col_idx, value=field)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_offset, item in enumerate(payload['rows'], start=5):
            for col_idx, field in enumerate(fieldnames, start=1):
                value = item.get(field)
                sheet.cell(row=row_offset, column=col_idx, value=value)

        for col_idx, field in enumerate(fieldnames, start=1):
            max_len = max(len(str(field)), 12)
            for item in payload['rows'][:500]:
                max_len = max(max_len, min(50, len(str(item.get(field, '')))))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        sheet.freeze_panes = 'A5'
        end_row = max(4, len(payload['rows']) + 4)
        end_col = get_column_letter(max(1, len(fieldnames)))
        sheet.auto_filter.ref = f'A4:{end_col}{end_row}'

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    # validation: ensure workbook opens cleanly
    checker = io.BytesIO(mem.getvalue())
    load_workbook(checker, read_only=True)

    return mem, f'scanoil-backup-{timestamp}.xlsx'


def build_backup_zip(include_files: bool = True, timestamp: str | None = None, include_excel: bool = True):
    timestamp = timestamp or datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    data = collect_backup_rows()
    mem = io.BytesIO()
    manifest = {
        'created_at_utc': datetime.utcnow().isoformat() + 'Z',
        'include_files': bool(include_files),
        'include_excel': bool(include_excel),
        'files_exported': 0,
        'files_failed': [],
        'tables': {folder: len(payload['rows']) for folder, payload in data.items()},
    }

    with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as zf:
        for folder, payload in data.items():
            rows = payload['rows']
            fieldnames = payload['fieldnames']
            zf.writestr(f'db/{folder}.json', json.dumps(rows, ensure_ascii=False, indent=2))
            zf.writestr(f'db/{folder}.csv', _csv_bytes(rows, fieldnames))

        if include_excel:
            excel_mem, excel_filename = build_backup_excel_bytes(data=data, timestamp=timestamp)
            zf.writestr(f'excel/{excel_filename}', excel_mem.getvalue())
            manifest['excel_file'] = f'excel/{excel_filename}'

        if include_files:
            sb = _get_supabase_client()
            if sb is not None:
                for row in ContractFile.query.order_by(ContractFile.id.asc()).all():
                    key = row.storage_key or row.storage_path
                    if not key:
                        manifest['files_failed'].append({'id': row.id, 'reason': 'missing_key'})
                        continue
                    bucket = row.bucket or 'contracts'
                    safe_name = row.original_name or f'{row.id}.pdf'
                    safe_name = os.path.basename(safe_name).replace('/', '_')
                    arcname = f'files/contracts/{row.contract_id}/{row.id}_{safe_name}'
                    try:
                        payload = sb.storage.from_(bucket).download(key)
                        if payload:
                            zf.writestr(arcname, payload)
                            manifest['files_exported'] += 1
                        else:
                            manifest['files_failed'].append({'id': row.id, 'reason': 'empty_payload'})
                    except Exception as exc:
                        current_app.logger.exception('Backup file export failed for ContractFile id=%s', row.id)
                        manifest['files_failed'].append({'id': row.id, 'reason': str(exc)[:200]})
            else:
                manifest['files_failed'].append({'id': None, 'reason': 'SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set'})

        zf.writestr('manifest.json', json.dumps(manifest, ensure_ascii=False, indent=2))

    mem.seek(0)
    return mem, f'scanoil-backup-{timestamp}.zip', manifest


def _build_cloud_key(filename: str, base_path: str = 'auto') -> str:
    stamp = datetime.utcnow()
    parts = []
    if base_path:
        parts.append(base_path.strip('/ '))
    parts.extend([str(stamp.year), f'{stamp.month:02d}', filename])
    return posixpath.join(*parts)


def _normalize_storage_list(items, prefix=''):
    normalized = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        name = item.get('name') or ''
        if not name:
            continue
        current = posixpath.join(prefix, name) if prefix else name
        metadata = item.get('metadata') or {}
        if metadata:
            normalized.append(current)
            continue
        normalized.extend(_normalize_storage_list(item.get('items') or [], current))
    return normalized


def cleanup_old_cloud_backups(sb, bucket: str, keep_last: int = 30, root_path: str = 'auto', suffixes=('.zip',)):
    keep_last = max(1, int(keep_last or 1))
    root_path = (root_path or '').strip('/ ')
    month_prefix = root_path
    suffixes = tuple(s.lower() for s in (suffixes or ('.zip',)))
    try:
        listing = sb.storage.from_(bucket).list(month_prefix or None, {'limit': 1000, 'sortBy': {'column': 'name', 'order': 'desc'}})
        all_keys = _normalize_storage_list(listing, month_prefix)
        filtered_keys = sorted([k for k in all_keys if k.lower().endswith(suffixes)], reverse=True)
        stale = filtered_keys[keep_last:]
        if stale:
            sb.storage.from_(bucket).remove(stale)
        return {'removed': len(stale), 'kept': min(len(filtered_keys), keep_last)}
    except Exception as exc:
        if current_app:
            current_app.logger.warning('Cloud backup cleanup failed: %s', exc)
        return {'removed': 0, 'kept': 0, 'error': str(exc)[:200]}


def upload_backup_bytes_to_supabase(bundle_bytes: bytes, filename: str, bucket: str = 'backups', base_path: str = 'auto', keep_last: int = 30, content_type: str | None = None):
    sb = _get_supabase_client()
    if sb is None:
        raise RuntimeError('SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set')

    key = _build_cloud_key(filename, base_path=base_path)
    content_type = content_type or (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        if filename.lower().endswith('.xlsx') else 'application/zip'
    )
    sb.storage.from_(bucket).upload(
        path=key,
        file=bundle_bytes,
        file_options={'content-type': content_type, 'cache-control': '3600', 'upsert': 'false'}
    )
    cleanup = cleanup_old_cloud_backups(
        sb,
        bucket=bucket,
        keep_last=keep_last,
        root_path=base_path,
        suffixes=(os.path.splitext(filename)[1] or '',),
    )
    return {'bucket': bucket, 'key': key, 'cleanup': cleanup}
